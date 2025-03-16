import os
import io
import requests
import openpyxl
from openpyxl import Workbook
from flask import (Flask, render_template, request, jsonify,
                   redirect, url_for, flash, send_file, session as flask_session)
from sqlalchemy import create_engine
from sqlalchemy.orm import scoped_session, sessionmaker

from db.config import DB_URI, API_TOKEN, BASE_API_URL, API_EMAIL, API_PASSWORD
from db.models import Base, Trip

app = Flask(__name__)
app.secret_key = "your_secret_key"  # for flashing and session

# Create engine with SQLite thread-safety; disable expire_on_commit so instances remain populated.
engine = create_engine(DB_URI, echo=True, connect_args={"check_same_thread": False})
Session = scoped_session(sessionmaker(bind=engine, expire_on_commit=False))

@app.teardown_appcontext
def shutdown_session(exception=None):
    Session.remove()

# --- Save Filter utilities using Flask session ---
def get_saved_filters():
    return flask_session.get("saved_filters", {})

def save_filter_to_session(name, filters):
    saved = flask_session.get("saved_filters", {})
    saved[name] = filters
    flask_session["saved_filters"] = saved

# ---------------------------
# Carrier grouping
# ---------------------------
CARRIER_GROUPS = {
    "Vodafone": ["vodafone", "voda fone", "tegi ne3eesh"],
    "Orange": ["orange", "orangeeg", "orange eg"],
    "Etisalat": ["etisalat", "e& etisalat", "e&"],
    "We": ["we"]
}

def normalize_carrier(carrier_name):
    if not carrier_name:
        return ""
    lower = carrier_name.lower().strip()
    for group, variants in CARRIER_GROUPS.items():
        for variant in variants:
            if variant in lower:
                return group
    return carrier_name.title()

# ---------------------------
# Token utilities
# ---------------------------
def fetch_api_token():
    url = f"{BASE_API_URL}/auth/sign_in"
    payload = {"admin_user": {"email": API_EMAIL, "password": API_PASSWORD}}
    resp = requests.post(url, json=payload)
    if resp.status_code == 200:
        return resp.json().get("token", None)
    else:
        print("Error fetching primary token:", resp.text)
        return None

def fetch_api_token_alternative():
    alt_email = "SupplyPartner@illa.com.eg"
    alt_password = "654321"
    url = f"{BASE_API_URL}/auth/sign_in"
    payload = {"admin_user": {"email": alt_email, "password": alt_password}}
    try:
        resp = requests.post(url, json=payload)
        resp.raise_for_status()
        return resp.json().get("token", None)
    except Exception as e:
        print("Error fetching alternative token:", e)
        return None

# ---------------------------
# Load Excel data
# ---------------------------
def load_excel_data(excel_path):
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    headers = []
    data = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            headers = row
        else:
            row_dict = {headers[j]: row[j] for j in range(len(row))}
            data.append(row_dict)
    print(f"Loaded {len(data)} rows from Excel.")
    return data

# ---------------------------
# Fetch trip data from API with fallback
# ---------------------------
def fetch_trip_from_api(trip_id, token=API_TOKEN):
    url = f"{BASE_API_URL}/trips/{trip_id}"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    try:
        resp = requests.get(url, headers=headers)
        resp.raise_for_status()
        data = resp.json()
        calc = data.get("data", {}).get("attributes", {}).get("calculatedDistance")
        if not calc or calc in [None, "", "N/A"]:
            raise ValueError("Missing calculatedDistance")
        return data
    except Exception as e:
        print("Error fetching trip data with primary token:", e)
        alt_token = fetch_api_token_alternative()
        if alt_token:
            headers = {"Authorization": f"Bearer {alt_token}", "Content-Type": "application/json"}
            try:
                resp = requests.get(url, headers=headers)
                resp.raise_for_status()
                data = resp.json()
                data["used_alternative"] = True
                return data
            except Exception as e:
                print("Error fetching trip data with alternative token:", e)
                return None
        else:
            return None

# ---------------------------
# Update or create Trip DB record
# ---------------------------
def update_trip_db(trip_id):
    session_local = Session()
    try:
        db_trip = session_local.query(Trip).filter_by(trip_id=trip_id).first()
        if db_trip and db_trip.status and db_trip.status.lower() == "completed":
            return db_trip
        api_data = fetch_trip_from_api(trip_id)
        if api_data and "data" in api_data:
            trip_attributes = api_data["data"]["attributes"]
            if db_trip is None:
                db_trip = Trip(
                    trip_id=trip_id,
                    status=trip_attributes.get("status"),
                    manual_distance=trip_attributes.get("manualDistance"),
                    calculated_distance=trip_attributes.get("calculatedDistance")
                )
                session_local.add(db_trip)
            else:
                db_trip.status = trip_attributes.get("status")
                try:
                    db_trip.manual_distance = float(trip_attributes.get("manualDistance") or 0)
                except ValueError:
                    db_trip.manual_distance = None
                try:
                    db_trip.calculated_distance = float(trip_attributes.get("calculatedDistance") or 0)
                except ValueError:
                    db_trip.calculated_distance = None
            if api_data.get("used_alternative"):
                db_trip.supply_partner = True
            session_local.commit()
        return db_trip
    except Exception as e:
        print("Error in update_trip_db:", e)
        session_local.rollback()
        return session_local.query(Trip).filter_by(trip_id=trip_id).first()
    finally:
        session_local.close()

# ---------------------------
# Route: Update Entire Database from Excel
# ---------------------------
@app.route("/update_db", methods=["POST"])
def update_db():
    session_local = Session()
    excel_path = os.path.join("data", "data.xlsx")
    excel_data = load_excel_data(excel_path)
    updated_ids = []
    for row in excel_data:
        trip_id = row.get("tripId")
        if trip_id:
            db_trip = update_trip_db(trip_id)
            if db_trip:
                updated_ids.append(trip_id)
    session_local.close()
    flash(f"Updated database for {len(updated_ids)} trips.", "success")
    return redirect(url_for("trips"))

# ---------------------------
# Route: Export filtered trips as XLSX
# ---------------------------
@app.route("/export_trips")
def export_trips():
    session_local = Session()
    # Get filter parameters from request
    filters = {
        "driver": request.args.get("driver"),
        "trip_id": request.args.get("trip_id"),
        "route_quality": request.args.get("route_quality"),
        "model": request.args.get("model"),
        "ram": request.args.get("ram"),
        "carrier": request.args.get("carrier"),
        "variance_min": request.args.get("variance_min"),
        "variance_max": request.args.get("variance_max"),
        "export_name": request.args.get("export_name", "exported_trips")
    }
    excel_path = os.path.join("data", "data.xlsx")
    excel_data = load_excel_data(excel_path)
    if filters["driver"]:
        excel_data = [row for row in excel_data if str(row.get("UserName", "")).strip() == filters["driver"]]
    if filters["trip_id"]:
        try:
            tid = int(filters["trip_id"])
            excel_data = [row for row in excel_data if row.get("tripId") == tid]
        except ValueError:
            pass
    if filters["model"]:
        excel_data = [row for row in excel_data if str(row.get("model", "")).strip() == filters["model"]]
    if filters["ram"]:
        excel_data = [row for row in excel_data if str(row.get("RAM", "")).strip() == filters["ram"]]
    if filters["carrier"]:
        excel_data = [row for row in excel_data if str(row.get("carrier", "")).strip().lower() == filters["carrier"].lower()]
    if filters["route_quality"]:
        excel_data = [row for row in excel_data if str(row.get("route_quality", "")) == filters["route_quality"]]
        
    # Batch merge with DB data
    excel_trip_ids = [row.get("tripId") for row in excel_data if row.get("tripId")]
    db_trips = session_local.query(Trip).filter(Trip.trip_id.in_(excel_trip_ids)).all()
    db_trip_map = {trip.trip_id: trip for trip in db_trips}
    merged = []
    for row in excel_data:
        trip_id = row.get("tripId")
        db_trip = db_trip_map.get(trip_id)
        if db_trip:
            try:
                md = float(db_trip.manual_distance)
            except (TypeError, ValueError):
                md = None
            try:
                cd = float(db_trip.calculated_distance)
            except (TypeError, ValueError):
                cd = None
            row["route_quality"] = db_trip.route_quality or ""
            row["manual_distance"] = md if md is not None else ""
            row["calculated_distance"] = cd if cd is not None else ""
            if md and cd and md != 0:
                pct = (cd / md) * 100
                row["distance_percentage"] = f"{pct:.2f}%"
                variance = abs(cd - md) / md * 100
                row["variance"] = variance
            else:
                row["distance_percentage"] = "N/A"
                row["variance"] = None
        else:
            row["route_quality"] = ""
            row["manual_distance"] = ""
            row["calculated_distance"] = ""
            row["distance_percentage"] = "N/A"
            row["variance"] = None
        merged.append(row)
    
    # Apply variance filters if provided
    if filters["variance_min"]:
        try:
            vmin = float(filters["variance_min"])
            merged = [r for r in merged if r.get("variance") is not None and r["variance"] >= vmin]
        except ValueError:
            pass
    if filters["variance_max"]:
        try:
            vmax = float(filters["variance_max"])
            merged = [r for r in merged if r.get("variance") is not None and r["variance"] <= vmax]
        except ValueError:
            pass

    wb = Workbook()
    ws = wb.active
    if merged:
        headers = list(merged[0].keys())
        ws.append(headers)
        for row in merged:
            ws.append([row.get(col) for col in headers])
    else:
        ws.append(["No data found"])
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    filename = f"{filters['export_name']}.xlsx"
    session_local.close()
    return send_file(file_stream, as_attachment=True, attachment_filename=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ---------------------------
# Dashboard (Analytics)
# ---------------------------
@app.route("/")
def analytics():
    session_local = Session()
    driver_filter = request.args.get("driver")
    excel_path = os.path.join("data", "data.xlsx")
    excel_data = load_excel_data(excel_path)
    # Build full driver list for dropdown
    drivers = sorted({str(row.get("UserName", "")).strip() for row in excel_data if row.get("UserName")})
    
    carrier_counts = {}
    os_counts = {}
    manufacturer_counts = {}
    model_counts = {}
    for row in excel_data:
        carrier = row.get("carrier")
        if carrier:
            norm = normalize_carrier(carrier)
            carrier_counts[norm] = carrier_counts.get(norm, 0) + 1
        os_ver = row.get("Android Version", "Unknown")
        os_counts[os_ver] = os_counts.get(os_ver, 0) + 1
        manufacturer = row.get("manufacturer", "Unknown")
        manufacturer_counts[manufacturer] = manufacturer_counts.get(manufacturer, 0) + 1
        model = row.get("model", "Unknown")
        model_counts[model] = model_counts.get(model, 0) + 1

    total_users = len(excel_data)
    device_usage = []
    for model, count in model_counts.items():
        percentage = (count / total_users) * 100 if total_users > 0 else 0
        device_usage.append({"model": model, "count": count, "percentage": round(percentage, 2)})

    trips_db = session_local.query(Trip).all()
    correct = 0
    incorrect = 0
    for trip in trips_db:
        try:
            md = float(trip.manual_distance)
            cd = float(trip.calculated_distance)
        except (TypeError, ValueError):
            continue
        if md and md != 0:
            if abs(cd - md) / md <= 0.2:
                correct += 1
            else:
                incorrect += 1
    total_trips = correct + incorrect
    if total_trips > 0:
        correct_pct = correct / total_trips * 100
        incorrect_pct = incorrect / total_trips * 100
    else:
        correct_pct = incorrect_pct = 0

    session_local.close()
    return render_template("analytics.html",
                           driver_filter=driver_filter,
                           drivers=drivers,
                           carrier_counts=carrier_counts,
                           os_counts=os_counts,
                           manufacturer_counts=manufacturer_counts,
                           device_usage=device_usage,
                           total_trips=total_trips,
                           correct_pct=correct_pct,
                           incorrect_pct=incorrect_pct)

# ---------------------------
# Trips Page with Variance, Pagination, and Batch DB Merge
# ---------------------------
@app.route("/trips")
def trips():
    session_local = Session()
    page = request.args.get("page", type=int, default=1)
    page_size = 100
    if page < 1:
        page = 1

    driver_filter = request.args.get("driver")
    trip_id_search = request.args.get("trip_id")
    route_quality_filter = request.args.get("route_quality")
    model_filter = request.args.get("model")
    ram_filter = request.args.get("ram")
    carrier_filter = request.args.get("carrier")
    variance_min = request.args.get("variance_min", type=float)
    variance_max = request.args.get("variance_max", type=float)

    excel_path = os.path.join("data", "data.xlsx")
    excel_data = load_excel_data(excel_path)

    if driver_filter:
        excel_data = [row for row in excel_data if str(row.get("UserName", "")).strip() == driver_filter]
    if trip_id_search:
        try:
            tid = int(trip_id_search)
            excel_data = [row for row in excel_data if row.get("tripId") == tid]
        except ValueError:
            pass
    if model_filter:
        excel_data = [row for row in excel_data if str(row.get("model", "")).strip() == model_filter]
    if ram_filter:
        excel_data = [row for row in excel_data if str(row.get("RAM", "")).strip() == ram_filter]
    if carrier_filter:
        excel_data = [row for row in excel_data if str(row.get("carrier", "")).strip().lower() == carrier_filter.lower()]

    # For filter dropdowns, compute full driver list from Excel
    full_excel = load_excel_data(excel_path)
    drivers = sorted({str(row.get("UserName", "")).strip() for row in full_excel if row.get("UserName")})

    # Batch query DB for the filtered Excel trip IDs.
    excel_trip_ids = [row.get("tripId") for row in excel_data if row.get("tripId")]
    db_trips = session_local.query(Trip).filter(Trip.trip_id.in_(excel_trip_ids)).all()
    db_trip_map = {trip.trip_id: trip for trip in db_trips}

    for row in excel_data:
        trip_id = row.get("tripId")
        db_trip = db_trip_map.get(trip_id)
        if db_trip:
            try:
                md = float(db_trip.manual_distance)
            except (TypeError, ValueError):
                md = None
            try:
                cd = float(db_trip.calculated_distance)
            except (TypeError, ValueError):
                cd = None
            row["route_quality"] = db_trip.route_quality or ""
            row["manual_distance"] = md if md is not None else ""
            row["calculated_distance"] = cd if cd is not None else ""
            if md and cd and md != 0:
                pct = (cd / md) * 100
                row["distance_percentage"] = f"{pct:.2f}%"
                variance = abs(cd - md) / md * 100
                row["variance"] = variance
            else:
                row["distance_percentage"] = "N/A"
                row["variance"] = None
        else:
            row["route_quality"] = ""
            row["manual_distance"] = ""
            row["calculated_distance"] = ""
            row["distance_percentage"] = "N/A"
            row["variance"] = None

    if route_quality_filter:
        excel_data = [row for row in excel_data if row.get("route_quality", "") == route_quality_filter]
    if variance_min is not None:
        excel_data = [row for row in excel_data if row.get("variance") is not None and row["variance"] >= variance_min]
    if variance_max is not None:
        excel_data = [row for row in excel_data if row.get("variance") is not None and row["variance"] <= variance_max]

    total_rows = len(excel_data)
    total_pages = (total_rows + page_size - 1) // page_size if total_rows else 1
    if page > total_pages and total_pages > 0:
        page = total_pages
    start_index = (page - 1) * page_size
    end_index = start_index + page_size
    page_data = excel_data[start_index:end_index]

    session_local.close()
    return render_template("trips.html",
                           driver_filter=driver_filter,
                           trips=page_data,
                           trip_id_search=trip_id_search or "",
                           route_quality_filter=route_quality_filter or "",
                           model_filter=model_filter or "",
                           ram_filter=ram_filter or "",
                           carrier_filter=carrier_filter or "",
                           variance_min=variance_min if variance_min is not None else "",
                           variance_max=variance_max if variance_max is not None else "",
                           total_rows=total_rows,
                           page=page,
                           total_pages=total_pages,
                           page_size=page_size)

# ---------------------------
# Trip Detail Page
# ---------------------------
@app.route("/trip/<int:trip_id>")
def trip_detail(trip_id):
    session_local = Session()
    db_trip = update_trip_db(trip_id)
    if db_trip and db_trip.status and db_trip.status.lower() == "completed":
        api_data = None
    else:
        api_data = fetch_trip_from_api(trip_id)
    trip_attributes = {}
    if api_data and "data" in api_data:
        trip_attributes = api_data["data"]["attributes"]

    excel_path = os.path.join("data", "data.xlsx")
    excel_data = load_excel_data(excel_path)
    excel_trip_data = None
    for row in excel_data:
        if row.get("tripId") == trip_id:
            excel_trip_data = row
            break

    distance_verification = "N/A"
    trip_insight = ""
    distance_percentage = "N/A"
    if db_trip:
        try:
            md = float(db_trip.manual_distance)
        except (TypeError, ValueError):
            md = None
        try:
            cd = float(db_trip.calculated_distance)
        except (TypeError, ValueError):
            cd = None
        if md is not None and cd is not None:
            lower_bound = md * 0.8
            upper_bound = md * 1.2
            if lower_bound <= cd <= upper_bound:
                distance_verification = "Calculated distance is true"
                trip_insight = "Trip data is consistent."
            else:
                distance_verification = "Manual distance is true"
                trip_insight = "Trip data is inconsistent."
            if md != 0:
                distance_percentage = f"{(cd / md * 100):.2f}%"
        else:
            distance_verification = "N/A"
            trip_insight = "N/A"
            distance_percentage = "N/A"

    session_local.close()
    return render_template("trip_detail.html",
                           db_trip=db_trip,
                           trip_attributes=trip_attributes,
                           excel_trip_data=excel_trip_data,
                           distance_verification=distance_verification,
                           trip_insight=trip_insight,
                           distance_percentage=distance_percentage)

# ---------------------------
# Update Route Quality (AJAX) - Auto on change
# ---------------------------
@app.route("/update_route_quality", methods=["POST"])
def update_route_quality():
    session_local = Session()
    data = request.get_json()
    trip_id = data.get("trip_id")
    quality = data.get("route_quality")
    db_trip = session_local.query(Trip).filter_by(trip_id=trip_id).first()
    if not db_trip:
        db_trip = Trip(
            trip_id=trip_id,
            route_quality=quality,
            status="",
            manual_distance=None,
            calculated_distance=None
        )
        session_local.add(db_trip)
    else:
        db_trip.route_quality = quality
    session_local.commit()
    session_local.close()
    return jsonify({"status": "success", "message": "Route quality updated."}), 200

# ---------------------------
# Trip Insights Page
# ---------------------------
@app.route("/trip_insights")
def trip_insights():
    session_local = Session()
    trips_db = session_local.query(Trip).all()
    quality_counts = {"Low": 0, "Moderate": 0, "High": 0, "": 0}
    total_manual = 0
    total_calculated = 0
    count_manual = 0
    count_calculated = 0
    consistent = 0
    inconsistent = 0
    for trip in trips_db:
        quality = trip.route_quality if trip.route_quality else ""
        quality_counts[quality] = quality_counts.get(quality, 0) + 1
        try:
            md = float(trip.manual_distance)
            cd = float(trip.calculated_distance)
        except (TypeError, ValueError):
            continue
        total_manual += md
        total_calculated += cd
        count_manual += 1
        count_calculated += 1
        if md != 0 and abs(cd - md) / md <= 0.2:
            correct = 1
            consistent += 1
        else:
            incorrect = 1
            inconsistent += 1
    avg_manual = total_manual / count_manual if count_manual > 0 else 0
    avg_calculated = total_calculated / count_calculated if count_calculated > 0 else 0
    session_local.close()
    return render_template("trip_insights.html",
                           quality_counts=quality_counts,
                           avg_manual=avg_manual,
                           avg_calculated=avg_calculated,
                           consistent=consistent,
                           inconsistent=inconsistent)

# ---------------------------
# Save Filter (store current filter parameters)
# ---------------------------
@app.route("/save_filter", methods=["POST"])
def save_filter():
    # Get current filter parameters from form data and a filter name
    filter_name = request.form.get("filter_name")
    filters = {
        "trip_id": request.form.get("trip_id"),
        "route_quality": request.form.get("route_quality"),
        "model": request.form.get("model"),
        "ram": request.form.get("ram"),
        "carrier": request.form.get("carrier"),
        "variance_min": request.form.get("variance_min"),
        "variance_max": request.form.get("variance_max"),
        "driver": request.form.get("driver")
    }
    if filter_name:
        saved = flask_session.get("saved_filters", {})
        saved[filter_name] = filters
        flask_session["saved_filters"] = saved
        flash(f"Filter '{filter_name}' saved.", "success")
    else:
        flash("Please provide a filter name.", "danger")
    return redirect(url_for("trips"))

# ---------------------------
# Apply Saved Filter (redirect to /trips with the saved filter parameters)
# ---------------------------
@app.route("/apply_filter/<filter_name>")
def apply_filter(filter_name):
    saved = flask_session.get("saved_filters", {})
    filters = saved.get(filter_name)
    if filters:
        # Build a query string from the saved filters.
        qs = "&".join(f"{key}={value}" for key, value in filters.items() if value)
        return redirect(url_for("trips") + "?" + qs)
    else:
        flash("Saved filter not found.", "danger")
        return redirect(url_for("trips"))

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
